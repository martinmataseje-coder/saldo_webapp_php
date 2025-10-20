# saldo_core.py
from io import BytesIO
from typing import Literal, Optional
import datetime as _dt
import unicodedata  # <- robustné porovnávanie textu

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# PDF export
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

HEADER_ROW = 9
DATE_FMT   = "DD.MM.YY"

# ---------- helpers (xlsx) ----------
def _norm(s):
    """Normalizácia stringu: odstráni NBSP, diakritiku, oreže medzery a zníži na lower()."""
    if s is None:
        return ""
    s = str(s).replace("\u00A0", " ")  # NBSP -> space
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.strip().lower()

def _find_col(headers, name):
    target = _norm(name)
    for i, h in enumerate(headers, start=1):
        if _norm(h) == target:
            return i
    return None

def _last_data_row(ws, key_col):
    last = HEADER_ROW
    for r in range(HEADER_ROW+1, ws.max_row+1):
        if ws.cell(row=r, column=key_col).value not in (None, ""):
            last = r
    return last

def _style_ws(ws, c_doc, c_inv, c_dz, c_du, c_sn, c_typ, c_amt, c_bal, last, theme="blue"):
    # jemná 4ka téma pre XLSX (nie úplne rovnaké farby ako PDF, ale decentné)
    header_fill = PatternFill("solid", fgColor="EAFBF9")  # bledý tyrkys
    zebra_fill  = PatternFill("solid", fgColor="F7FDFB")
    head_font   = Font(bold=True, color="0F172A")
    thin = Side(style="thin", color="D0D7E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # hlavička (zapni zalamovanie textu)
    for c in range(1, ws.max_column+1):
        cell = ws.cell(row=HEADER_ROW, column=c)
        cell.font = head_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = border

    widths = {c_doc:16, c_inv:18, c_dz:18, c_du:16, c_sn:16, c_typ:22, c_amt:14, c_bal:14}
    for col_idx, w in widths.items():
        if col_idx:
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    # formát peňazí
    for r in range(HEADER_ROW+1, last+1):
        if c_amt: ws.cell(row=r, column=c_amt).number_format = '#,##0.00'
        if c_bal: ws.cell(row=r, column=c_bal).number_format = '#,##0.00'

    # zebra + okraje
    for r in range(HEADER_ROW+1, last+1):
        if (r - (HEADER_ROW+1)) % 2 == 0:
            for c in range(1, ws.max_column+1):
                ws.cell(row=r, column=c).fill = zebra_fill
                ws.cell(row=r, column=c).border = border

def _insert_logo_xlsx(ws, logo_bytes: Optional[bytes]):
    if not logo_bytes:
        return
    try:
        bio = BytesIO(logo_bytes)
        img = XLImage(bio)
        ws.add_image(img, "A1")
    except Exception:
        pass

# ---------- helpers (PDF) ----------
def _register_fonts():
    """Registruje DejaVu Sans (ak je v data/) a nastaví family mapovanie; inak padá na Helvetica."""
    try:
        import os
        from reportlab.lib.fonts import addMapping
        base = os.path.dirname(__file__)
        ttf_regular = os.path.join(base, "data", "DejaVuSans.ttf")
        ttf_bold    = os.path.join(base, "data", "DejaVuSans-Bold.ttf")
        if os.path.exists(ttf_regular) and os.path.exists(ttf_bold):
            pdfmetrics.registerFont(TTFont("DejaVuSans", ttf_regular))
            pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", ttf_bold))
            addMapping("DejaVuSans", 0, 0, "DejaVuSans")
            addMapping("DejaVuSans", 1, 0, "DejaVuSans-Bold")
            return ("DejaVuSans", "DejaVuSans-Bold")
    except Exception:
        pass
    return ("Helvetica", "Helvetica-Bold")

def _fmt_date(v):
    import datetime
    if isinstance(v, (datetime.datetime, _dt.datetime, datetime.date, _dt.date)):
        return v.strftime("%d.%m.%Y")
    s = str(v).strip() if v is not None else ""
    if " " in s: s = s.split(" ")[0]
    if "-" in s:
        parts = s.split("-")
        if len(parts) == 3 and all(parts):
            yyyy, mm, dd = parts
            return f"{dd}.{mm}.{yyyy}"
    return s

def _num(v):
    try:
        return float(v)
    except Exception:
        return None

def _fmt_money(x):
    if x is None:
        return ""
    s = f"{x:,.2f}".replace(",", " ")
    return s + "\u00A0€"

# Palety tém (pre PDF)
THEMES = {
    "blue": {"header_hex": "#25B3AD", "alt_row": "#F9FEFD", "grid": "#E2E8F0"},
    "gray": {"header_hex": "#4A5568", "alt_row": "#F7F7F7", "grid": "#D9D9D9"},
    "warm": {"header_hex": "#C6A875", "alt_row": "#FFF9F2", "grid": "#EADDC8"},
}

def _build_pdf(ws, hdr_meno, hdr_sap, hdr_ucet, hdr_spol, logo_bytes: Optional[bytes], theme="blue"):
    FONT_REG, FONT_BOLD = _register_fonts()
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="HdrTitle", parent=styles["Title"], fontName=FONT_BOLD, alignment=0))
    styles.add(ParagraphStyle(name="Base", parent=styles["Normal"], fontName=FONT_REG, fontSize=9, leading=12))
    styles.add(ParagraphStyle(name="HdrSmall", parent=styles["Normal"], fontName=FONT_BOLD, fontSize=9, alignment=1))
    styles.add(ParagraphStyle(name="Cell", parent=styles["Normal"], fontName=FONT_REG, fontSize=8, leading=10))
    styles.add(ParagraphStyle(name="CellRight", parent=styles["Normal"], fontName=FONT_REG, fontSize=8, leading=10, alignment=2))

    th = THEMES.get(theme, THEMES["blue"])

    # hlavičky z Excelu
    xhdrs = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, ws.max_column+1)]
    c_doc = _find_col(xhdrs, "Číslo dokladu")
    c_inv = _find_col(xhdrs, "číslo Faktúry") or _find_col(xhdrs, "Číslo Faktúry")
    # akceptuj viac variantov (s/bez medzery a s/bez zalomenia)
    c_dz  = (_find_col(xhdrs, "Dátum vystavenia / Pripísania platby")
             or _find_col(xhdrs, "Dátum vystavenia/Pripísania platby")
             or _find_col(xhdrs, "Dátum vystavenia /\nPripísania platby")
             or _find_col(xhdrs, "Dátum zadania"))
    c_du  = _find_col(xhdrs, "Dátum účtovania")
    c_sn  = _find_col(xhdrs, "Splatnosť netto")
    c_typ = _find_col(xhdrs, "Typ dokladu")
    c_amt = _find_col(xhdrs, "Čiastka")
    c_bal = _find_col(xhdrs, "Zostatok")
    last  = _last_data_row(ws, c_doc)

    pdf_hdrs = [
        "Č. dokladu",
        "Č. faktúry",
        "Dátum vystavenia /\nPripísania platby",
        "Dátum účt.",
        "Splatnosť",
        "Typ dokladu",
        "Čiastka",
        "Zostatok",
    ]

    data = [[Paragraph(h, styles["HdrSmall"]) for h in pdf_hdrs]]
    run_bal = 0.0
    def _is_faktura(txt): return isinstance(txt, str) and _norm(txt) == _norm("Faktúra")

    for r in range(HEADER_ROW+1, last+1):
        doc = ws.cell(row=r, column=c_doc).value
        inv = ws.cell(row=r, column=c_inv).value
        dz  = ws.cell(row=r, column=c_dz).value
        du  = ws.cell(row=r, column=c_du).value
        sn  = ws.cell(row=r, column=c_sn).value
        typ = ws.cell(row=r, column=c_typ).value
        amt = _num(ws.cell(row=r, column=c_amt).value)
        add_amt = amt if amt is not None else 0.0
        run_bal += add_amt

        row = [
            Paragraph("" if doc is None else str(doc), styles["Cell"]),
            Paragraph("" if (inv is None or not _is_faktura(typ)) else str(inv), styles["Cell"]),
            Paragraph(_fmt_date(dz), styles["Cell"]),
            Paragraph(_fmt_date(du), styles["Cell"]),
            Paragraph(_fmt_date(sn) if _is_faktura(typ) else "", styles["Cell"]),
            Paragraph("" if typ is None else str(typ), styles["Cell"]),
            Paragraph(_fmt_money(amt), styles["CellRight"]),
            Paragraph(_fmt_money(run_bal), styles["CellRight"]),
        ]
        data.append(row)

    # "Súčet"
    total_row = [Paragraph("", styles["Cell"]) for _ in range(8)]
    total_row[5] = Paragraph("<b>Súčet</b>", styles["HdrSmall"])
    total_row[7] = Paragraph(f"<b>{_fmt_money(run_bal)}</b>", styles["CellRight"])
    data.append(total_row)

    # Layout
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)

    # Hlavička PDF
    title = Paragraph("Náhľad na fakturačný účet – saldo", styles["HdrTitle"])
    date_p = Paragraph(f"Dátum generovania: {_dt.datetime.now().strftime('%d.%m.%Y')}", styles["Base"])
    meta = Paragraph(
        f"SWAN a.s. — <b>Meno:</b> {hdr_meno} • <b>SAP ID:</b> {hdr_sap} • <b>Zmluvný účet:</b> {hdr_ucet}",
        styles["Base"]
    )

    header_tbl_data = []
    if logo_bytes:
        rlimg = RLImage(BytesIO(logo_bytes), width=60, height=60)
        header_tbl_data.append([rlimg, Spacer(10, 10), [title, date_p, meta]])
        col_head_widths = [60, 10, None]
    else:
        header_tbl_data.append(["", "", [title, date_p, meta]])
        col_head_widths = [0, 0, None]

    header_tbl = Table(header_tbl_data, colWidths=col_head_widths, hAlign="LEFT")
    header_tbl.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ("TOPPADDING", (0,0), (-1,-1), 0),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))

    story = [header_tbl, Spacer(1, 6)]

    # tabuľka
    col_widths = [75, 60, 70, 58, 58, 70, 62, 68]
    table = Table(data, repeatRows=1, colWidths=col_widths, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor(th["header_hex"])),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor(th["grid"])),
        ("ROWBACKGROUNDS", (0,1), (-1,-2), [colors.white, colors.HexColor(th["alt_row"])]),
        ("ALIGN", (2,1), (4,-2), "CENTER"),
        ("ALIGN", (6,1), (7,-2), "RIGHT"),
        ("ALIGN", (7,-1), (7,-1), "RIGHT"),
        ("FONTSIZE", (0,0), (-1,0), 9),
        ("FONTSIZE", (0,1), (-1,-1), 8),
        ("LEFTPADDING", (0,0), (-1,-1), 2),
        ("RIGHTPADDING", (0,0), (-1,-1), 2),
        ("TOPPADDING", (0,0), (-1,-1), 2),
        ("BOTTOMPADDING",(0,0), (-1,-1), 2),
        ("BACKGROUND", (0,-1), (-1,-1), colors.HexColor(th["header_hex"])),
        ("TEXTCOLOR",  (0,-1), (-1,-1), colors.white),
        ("FONTNAME",   (5,-1), (5,-1), FONT_BOLD),
        ("FONTNAME",   (7,-1), (7,-1), FONT_BOLD),
    ]))

    story.append(table)
    doc.build(story)
    buf.seek(0)
    return buf.read()

# ---------- public API ----------
def generate_saldo_document(
    template_bytes: bytes,
    helper_bytes: bytes,
    src1_bytes: bytes,
    src2_bytes: bytes,
    hdr_meno: str,
    hdr_sap: str,
    hdr_ucet: str,
    hdr_spol: str = "SWAN a.s.",
    theme: Literal["blue","gray","warm"] = "blue",
    logo_bytes: Optional[bytes] = None,
    output: Literal["xlsx","pdf"] = "xlsx",
) -> bytes:
    """
    Vygeneruje XLSX alebo PDF:
      - mapuje 'Označenie pôvodu' -> 'Typ dokladu' pomocou pomôcky,
      - doplní 'Číslo faktúry' z 'Doplnková referencia' (src2),
      - vypočíta bežiaci 'Zostatok',
      - vloží hlavičku B1..B4 a voliteľne logo,
      - pre PDF použije firemnú tabuľku a témy.
    """
    # --- TEMPLATE ---
    wb = load_workbook(BytesIO(template_bytes), data_only=False)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, ws.max_column+1)]

    c_doc = _find_col(headers, "Číslo dokladu")
    c_inv = _find_col(headers, "číslo Faktúry") or _find_col(headers, "Číslo Faktúry")
    # akceptuj viac variantov (s/bez medzery a s/bez zalomenia)
    c_dz  = (_find_col(headers, "Dátum vystavenia / Pripísania platby")
             or _find_col(headers, "Dátum vystavenia/Pripísania platby")
             or _find_col(headers, "Dátum vystavenia /\nPripísania platby")
             or _find_col(headers, "Dátum zadania"))
    c_du  = _find_col(headers, "Dátum účtovania")
    c_sn  = _find_col(headers, "Splatnosť netto")
    c_typ = _find_col(headers, "Typ dokladu")
    c_amt = _find_col(headers, "Čiastka")
    c_bal = _find_col(headers, "Zostatok")

    if None in (c_doc, c_inv, c_dz, c_du, c_sn, c_typ, c_amt, c_bal):
        # Diagnostická správa, aby bolo hneď jasné, čo chýba
        missing = []
        if c_doc is None: missing.append("Číslo dokladu")
        if c_inv is None: missing.append("Číslo Faktúry/číslo Faktúry")
        if c_dz  is None: missing.append("Dátum vystavenia / Pripísania platby / Dátum zadania")
        if c_du  is None: missing.append("Dátum účtovania")
        if c_sn  is None: missing.append("Splatnosť netto")
        if c_typ is None: missing.append("Typ dokladu")
        if c_amt is None: missing.append("Čiastka")
        if c_bal is None: missing.append("Zostatok")
        raise RuntimeError(f"V TEMPLATE chýba niektorý povinný stĺpec. Chýbajú: {', '.join(missing)}")

    # Premenuj hlavičku „Dátum zadania“ -> nový text (ak je to práve tento stĺpec)
    hdr_cell = ws.cell(row=HEADER_ROW, column=c_dz)
    if _norm(hdr_cell.value) in (_norm("Dátum zadania"),
                                 _norm("Dátum vystavenia/Pripísania platby"),
                                 _norm("Dátum vystavenia / Pripísania platby")):
        hdr_cell.value = "Dátum vystavenia / Pripísania platby"
        hdr_cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    # --- HELPER (pomôcka) ---
    wb_h = load_workbook(BytesIO(helper_bytes), data_only=True); ws_h = wb_h[wb_h.sheetnames[0]]
    hdr_h = [ws_h.cell(row=1, column=c).value for c in range(1, ws_h.max_column+1)]
    def idx_h(name):
        for i,h in enumerate(hdr_h, start=1):
            if isinstance(h,str) and h.strip()==name:
                return i
        return None
    h_src = idx_h("Označenie pôvodu"); h_dst = idx_h("Typ dokladu")
    if not h_src or not h_dst:
        raise RuntimeError("V pomôcke chýba 'Označenie pôvodu' alebo 'Typ dokladu'.")

    pom_map = {}
    for r in range(2, ws_h.max_row+1):
        s = ws_h.cell(row=r, column=h_src).value
        t = ws_h.cell(row=r, column=h_dst).value
        if isinstance(s,str) and s.strip()!="":
            pom_map[s.strip()] = t.strip() if isinstance(t,str) else t

    # --- SRC1 (pohyby) + mapovanie typu ---
    wb1 = load_workbook(BytesIO(src1_bytes), data_only=True); ws1 = wb1[wb1.sheetnames[0]]
    hdr1 = [ws1.cell(row=1, column=c).value for c in range(1, ws1.max_column+1)]
    def idx1(name):
        for i,h in enumerate(hdr1, start=1):
            if isinstance(h,str) and h.strip()==name:
                return i
        return None
    i_doc = idx1("Číslo dokladu"); i_dz=idx1("Dátum zadania"); i_du=idx1("Dátum účtovania")
    i_sn  = idx1("Splatnosť netto"); i_op=idx1("Označenie pôvodu"); i_amt=idx1("Čiastka")

    # vyčisti dáta v šablóne (ponechaj hlavičku)
    if ws.max_row > HEADER_ROW:
        ws.delete_rows(HEADER_ROW+1, ws.max_row-HEADER_ROW)

    r0 = HEADER_ROW+1
    for r in range(2, ws1.max_row+1):
        row_has_data = any(ws1.cell(row=r, column=c).value not in (None,"") for c in range(1, ws1.max_column+1))
        if not row_has_data:
            continue
        ozn_pov = ws1.cell(row=r, column=i_op).value if i_op else None
        mapped_typ = pom_map.get(ozn_pov.strip() if isinstance(ozn_pov, str) else ozn_pov, None)

        # plnenie štandardných polí
        ws.cell(row=r0, column=c_doc, value=ws1.cell(row=r, column=i_doc).value if i_doc else None)
        ws.cell(row=r0, column=c_dz,  value=ws1.cell(row=r, column=i_dz).value if i_dz else None)
        ws.cell(row=r0, column=c_du,  value=ws1.cell(row=r, column=i_du).value if i_du else None)

        # Splatnosť len pri faktúrach, inak None
        if mapped_typ and isinstance(mapped_typ, str) and _norm(mapped_typ) == _norm("Faktúra"):
            ws.cell(row=r0, column=c_sn, value=ws1.cell(row=r, column=i_sn).value if i_sn else None)
        else:
            ws.cell(row=r0, column=c_sn, value=None)

        ws.cell(row=r0, column=c_typ, value=mapped_typ if mapped_typ is not None else None)
        ws.cell(row=r0, column=c_amt, value=ws1.cell(row=r, column=i_amt).value if i_amt else None)
        r0 += 1

    # --- Zostatok + formát dátumov ---
    L_G = get_column_letter(c_amt); L_H = get_column_letter(c_bal)
    last = _last_data_row(ws, c_doc)
    for r in range(HEADER_ROW+1, last+1):
        ws.cell(row=r, column=c_bal, value=f"={L_G}{r}" if r==HEADER_ROW+1 else f"={L_H}{r-1}+{L_G}{r}")
    for c in (c_dz, c_du, c_sn):
        if c:
            for rr in range(HEADER_ROW+1, last+1):
                ws.cell(row=rr, column=c).number_format = DATE_FMT

    # --- SRC2 (väzby) – doplň „Číslo faktúry“ z „Doplnková referencia“ ---
    wb2 = load_workbook(BytesIO(src2_bytes), data_only=True); ws2 = wb2[wb2.sheetnames[0]]
    hdr2 = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column+1)]
    def idx2(name):
        for i,h in enumerate(hdr2, start=1):
            if isinstance(h,str) and h.strip()==name:
                return i
        return None
    j_doc = idx2("Číslo dokladu"); j_ref = idx2("Doplnková referencia")
    if not j_doc or not j_ref:
        raise RuntimeError("V zdroji 2 chýba 'Číslo dokladu' alebo 'Doplnková referencia'.")

    ref_map = {}
    for r in range(2, ws2.max_row+1):
        k = ws2.cell(row=r, column=j_doc).value
        v = ws2.cell(row=r, column=j_ref).value
        if k not in (None,""):
            s = ""
            if isinstance(v, str):
                s = v.strip()
                if s.upper().startswith("VBRK"): s = s[4:].strip()
            elif v is not None:
                s = str(v)
            ref_map[str(k).strip()] = s

    def is_faktura(v): return isinstance(v, str) and _norm(v) == _norm("Faktúra")
    for rr in range(HEADER_ROW+1, last+1):
        doc = ws.cell(row=rr, column=c_doc).value
        typ = ws.cell(row=rr, column=c_typ).value
        if is_faktura(typ):
            inv = ref_map.get(str(doc).strip() if doc not in (None,"") else "", "")
            ws.cell(row=rr, column=c_inv, value=inv if inv else None)
        else:
            ws.cell(row=rr, column=c_inv, value=None)

    # --- horná hlavička pre XLSX + logo + štýl
    ws["B1"] = hdr_sap; ws["B2"] = hdr_meno; ws["B3"] = hdr_spol; ws["B4"] = hdr_ucet
    _insert_logo_xlsx(ws, logo_bytes)
    _style_ws(ws, c_doc, c_inv, c_dz, c_du, c_sn, c_typ, c_amt, c_bal, last, theme=theme)

    # --- export
    if output == "pdf":
        return _build_pdf(ws, hdr_meno, hdr_sap, hdr_ucet, hdr_spol, logo_bytes=logo_bytes, theme=theme)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
