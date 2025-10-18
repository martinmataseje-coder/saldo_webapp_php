# -*- coding: utf-8 -*-
"""
Saldo PDF layout v2.9 (final)
- A4 na výšku, logo kruh 4+KA (transparentné)
- Titul: "Náhľad na fakturačný účet – saldo" (bez bodky)
- Hlavička zákazníka: SAP ID · Meno zákazníka · Zmluvný účet · Názov spoločnosti
  (labels jemný bold: DejaVuSans-Bold, 8.5 pt, #333333; hodnoty 9 pt)
- Tabuľka:
  - žiadne prepočty, len zobrazovanie hodnôt z XLS
  - dátumy dd-mm-yy, "číslo Faktúry" ako text (odstránené 'VBRK'), "Čiastka" s €
  - stĺpec "Zostatok" má prioritu šírky (+20 %)
  - zarovnanie: čísla vpravo, dátumy stred, ostatné vľavo
- Pätička:
  - "Celkový zostatok: ..." = posledná neprázdna hodnota v stĺpci "Zostatok"
"""

from typing import Optional
import pandas as pd
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


def render_saldo_pdf(
    excel_path: str,
    logo_path: str,
    output_pdf: str,
    title_text: str = "Náhľad na fakturačný účet – saldo"
) -> None:
    """Vygeneruje PDF podľa layoutu v2.9 z už hotového XLS (bez prepočtov)."""

    def s(x) -> str:
        return "" if x is None else str(x)

    # Fonty (DejaVuSans má SK diakritiku)
    pdfmetrics.registerFont(TTFont("DejaVuSans", "DejaVuSans.ttf"))
    pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", "DejaVuSans-Bold.ttf"))
    FONT_REG = "DejaVuSans"
    FONT_BOLD = "DejaVuSans-Bold"

    # Načítanie XLS s computed values (žiadne prepočty tu nerobíme)
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    vals = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]

    # Nájdeme riadok hlavičky podľa "Číslo dokladu"
    header_idx: Optional[int] = None
    for i, row in enumerate(vals):
        line = " | ".join(s(x) for x in row).lower()
        if "číslo dokladu" in line or "cislo dokladu" in line:
            header_idx = i
            break
    if header_idx is None:
        # fallback: prvý nenulový riadok
        for i, row in enumerate(vals):
            if any(s(x).strip() for x in row):
                header_idx = i
                break

    header = [s(x) for x in vals[header_idx]]
    while header and header[-1] == "":
        header.pop()

    # Dáta pod hlavičkou
    rows = []
    for r in range(header_idx + 1, len(vals)):
        row_vals = [s(x) for x in vals[r][:len(header)]]
        if not any(v.strip() for v in row_vals):
            continue
        rows.append(row_vals)
    df = pd.DataFrame(rows, columns=header)

    # Hlavička zákazníka z prehlavičky (riadky nad db hlavičkou)
    cust = {"SAP ID": "", "Meno zákazníka": "", "Zmluvný účet": "", "Názov spoločnosti": ""}
    pre = vals[:header_idx]
    labels_map = {
        "sap": "SAP ID",
        "sap id": "SAP ID",
        "sapid": "SAP ID",
        "meno zákazníka": "Meno zákazníka",
        "meno zakaznika": "Meno zákazníka",
        "zmluvný účet": "Zmluvný účet",
        "zmluvny ucet": "Zmluvný účet",
        "názov spoločnosti": "Názov spoločnosti",
        "nazov spolocnosti": "Názov spoločnosti",
    }
    for row in pre:
        for idx, cell in enumerate(row):
            lab = s(cell).strip().lower()
            if not lab:
                continue
            for key, std in labels_map.items():
                if key in lab:
                    # hodnota = prvá neprázdna vpravo
                    val = ""
                    for v in row[idx + 1 :]:
                        sv = s(v).strip()
                        if sv:
                            val = sv
                            break
                    if val and not cust[std]:
                        cust[std] = val

    # Čistenie údajov (bez prepočtov)
    # "číslo Faktúry" ako text, vyčisti "VBRK"
    for candidate in ["číslo Faktúry", "číslo faktúry", "cislo faktury"]:
        if candidate in df.columns:
            cf = candidate
            df[cf] = df[cf].astype(str)
            df.loc[df[cf].str.contains("VBRK", case=False, na=False), cf] = ""
            df[cf] = df[cf].replace({"nan": "", "None": "", "NaN": ""})
            break

    # dátumy -> dd-mm-yy
    for c in df.columns:
        low = c.lower()
        if any(k in low for k in ["dátum", "datum", "splatnosť", "splatnost"]):
            parsed = pd.to_datetime(df[c], dayfirst=True, errors="coerce")
            out = []
            for orig, p in zip(df[c], parsed):
                if pd.isna(p):
                    out.append("" if str(orig).lower() in ("nan", "none", "nat", "") else str(orig))
                else:
                    out.append(p.strftime("%d-%m-%y"))
            df[c] = out

    # "Čiastka" s € (len formát, žiadne počítanie)
    if "Čiastka" in df.columns:
        def fmt_eur(v):
            t = s(v).strip()
            if t.lower() in ("nan", "none", ""):
                return ""
            norm = t.replace(" ", "").replace("\xa0", "").replace(",", ".")
            try:
                val = float(norm)
                return f"{val:,.2f}".replace(",", " ").replace(".", ",") + " €"
            except Exception:
                return t
        df["Čiastka"] = [fmt_eur(x) for x in df["Čiastka"]]

    # "Zostatok": len zobraz, nič neprepočítavaj
    if "Zostatok" in df.columns:
        df["Zostatok"] = df["Zostatok"].apply(
            lambda v: "" if s(v).strip().lower() in ("nan", "none", "nat", "") else s(v)
        )

    # ===== PDF ====
    PAGE_W, PAGE_H = A4
    LEFT = RIGHT = 10 * mm
    TOP = 16 * mm
    BOTTOM = 14 * mm
    CONTENT_W = PAGE_W - LEFT - RIGHT

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name="Title", fontName=FONT_BOLD, fontSize=16, leading=18)
    info_val = ParagraphStyle(name="InfoV", fontName=FONT_REG, fontSize=9, leading=12, textColor=colors.black)
    cell = ParagraphStyle(name="Cell", fontName=FONT_REG, fontSize=7.5, leading=9.5)
    cell_r = ParagraphStyle(name="CellR", fontName=FONT_REG, fontSize=7.5, leading=9.5, alignment=2)
    cell_c = ParagraphStyle(name="CellC", fontName=FONT_REG, fontSize=7.5, leading=9.5, alignment=1)
    head = ParagraphStyle(name="Head", fontName=FONT_BOLD, fontSize=8, leading=10, textColor=colors.white)

    # určenie typov pre zarovnanie
    num_cols = set([c for c in df.columns if c in ["Čiastka", "Zostatok"]])
    date_cols = set([c for c in df.columns if any(k in c.lower() for k in ["dátum", "datum", "splatnosť", "splatnost"])])

    # tabuľkové dáta
    table_data = []
    table_data.append([Paragraph(str(c), head) for c in df.columns])
    for _, r in df.iterrows():
        row_cells = []
        for c in df.columns:
            txt = "" if pd.isna(r[c]) else s(r[c])
            if c in num_cols:
                row_cells.append(Paragraph(txt, cell_r))
            elif c in date_cols:
                row_cells.append(Paragraph(txt, cell_c))
            else:
                row_cells.append(Paragraph(txt, cell))
        table_data.append(row_cells)

    # výpočet šírok stĺpcov (mix max/avg dĺžky)
    def col_weight(series, name):
        vals = [str(x) for x in series.head(300)] + [str(name)]
        lengths = [len(v) for v in vals if v]
        if not lengths:
            return 1
        max_len = max(lengths)
        avg_len = sum(lengths) / len(lengths)
        return 0.6 * max_len + 0.4 * avg_len

    weights = [col_weight(df[c], c) for c in df.columns]
    # +20 % priorita pre "Zostatok"
    if "Zostatok" in df.columns:
        z_idx = list(df.columns).index("Zostatok")
        weights[z_idx] *= 1.2

    tot = sum(weights) if sum(weights) > 0 else 1
    col_widths = [(w / tot) * CONTENT_W for w in weights]

    # dokument a hlavička
    doc = SimpleDocTemplate(
        output_pdf,
        pagesize=A4,
        leftMargin=LEFT,
        rightMargin=RIGHT,
        topMargin=TOP,
        bottomMargin=BOTTOM,
    )

    story = []

    # logo + titul
    logo_size_mm = 16.0
    logo = Image(logo_path, width=logo_size_mm * mm, height=logo_size_mm * mm, kind="proportional", mask="auto")
    title_para = Paragraph(title_text, title_style)
    header_tbl = Table([[logo, title_para]], colWidths=[logo_size_mm * mm + 4 * mm, CONTENT_W - (logo_size_mm * mm + 4 * mm)])
    header_tbl.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(header_tbl)
    story.append(Spacer(1, 4))

    # hlavička zákazníka (jemný bold na labeloch)
    parts = []
    order = [("SAP ID", cust.get("SAP ID", "")),
             ("Meno zákazníka", cust.get("Meno zákazníka", "")),
             ("Zmluvný účet", cust.get("Zmluvný účet", "")),
             ("Názov spoločnosti", cust.get("Názov spoločnosti", ""))]
    for label, value in order:
        if value:
            parts.append(
                f'<font name="{FONT_BOLD}" size="8.5" color="#333333">{label}:</font> '
                f'<font name="{FONT_REG}" size="9">{value}</font>'
            )
    if parts:
        story.append(Paragraph(" · ".join(parts), info_val))
        story.append(Spacer(1, 8))

    # tabuľka
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    header_bg = colors.HexColor("#BFEAF0")
    grid_color = colors.HexColor("#CFCFCF")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), header_bg),
                ("GRID", (0, 0), (-1, -1), 0.35, grid_color),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3.5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3.5),
                ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ]
        )
    )
    story.append(table)

    # pätička: posledná neprázdna hodnota v "Zostatok"
    last_nonempty = ""
    if "Zostatok" in df.columns:
        for val in reversed(df["Zostatok"].tolist()):
            s_val = str(val).strip()
            if s_val:
                last_nonempty = s_val
                break
    if last_nonempty:
        story.append(Spacer(1, 6))
        footer_style = ParagraphStyle(
            name="SaldoFooter",
            fontName=FONT_BOLD,
            fontSize=9,
            textColor=colors.black,
            alignment=2,  # right
        )
        story.append(Paragraph(f"Celkový zostatok: {last_nonempty}", footer_style))

    # export
    doc.build(story)
